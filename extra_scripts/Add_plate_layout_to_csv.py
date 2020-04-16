
import openpyxl
import os
import shutil
import numpy as np

peptide_layout = np.array([['No Pep', 'Pent'],
                           ['GRP22', 'Hex'],
                           ['GRP35', 'Hex2'],
                           ['GRP46', 'Hept'],
                           ['GRP51', '24D'],
                           ['GRP52', '24E'],
                           ['GRP63', '24K'],
                           ['GRP80', '17K']])

input_dir = input('Specify input directory:\n')
new_dir = '{}/Reformatted_csvs/'.format(input_dir)
if os.path.isdir(new_dir):
    shutil.rmtree(new_dir)
os.mkdir(new_dir)

repeat_conv = {'190417': 'repeat_1',
               '190424': 'repeat_2',
               '190429': 'repeat_3',
               '190510': 'repeat_4',
               '190515': 'repeat_5',
               '190517': 'repeat_6'}

for csv in os.listdir(input_dir):
    if csv.endswith('.xlsx') and not 'plate_layout' in csv.lower():
        date = csv.split('_')[0]
        csv_split = csv.split('_')[1:]
        analyte = '_'.join([section[0:1].upper() + section[1:].lower() for section in csv_split]).replace('.xlsx', '')

        plate_layout = np.array([['', '1:2', '3:4', '5:6', '7:8', '9:10', '11:12', '13:14', '15:16', '17:18', '19:20', '21:22', '23:24'],
                                 ['A:H', analyte, analyte, analyte, 'Blank', analyte, analyte, analyte, 'Blank', analyte, analyte, analyte, analyte],
                                 ['I:P', analyte, analyte, analyte, 'Blank', analyte, analyte, analyte, 'Blank', analyte, analyte, analyte, analyte]])

        new_csv = '{}/{}_{}.xlsx'.format(new_dir, analyte, repeat_conv[date])
        print('Saving {}'.format(new_csv))
        if os.path.isfile(new_csv):
            raise Exception('{} already exists'.format(new_csv))

        if date == '190429':
            orig_sheet_name = 'End point_1'
        else:
            orig_sheet_name = 'End point'
        orig_sheet = openpyxl.load_workbook(filename='{}/{}'.format(input_dir, csv))[orig_sheet_name]

        workbook = openpyxl.Workbook()
        # Copies End point worksheet from original csv file
        worksheet_1 = workbook.active
        worksheet_1.title = 'End point'
        for r in range (1, orig_sheet.max_row+1):
            for c in range (1, orig_sheet.max_column+1):
                cell = orig_sheet.cell(row=r, column=c)
                worksheet_1.cell(row=r, column=c).value = cell.value
        #Writes Protocol Information worksheet
        worksheet_2 = workbook.create_sheet(title='Protocol Information')
        worksheet_2.cell(row=2, column=1).value = 'Plate layout'
        for index, val in np.ndenumerate(plate_layout):
            r = index[0]
            c = index[1]
            worksheet_2.cell(row=r+3, column=c+1).value = val
        worksheet_2.cell(row=7, column=1).value = 'Peptide layout'
        for index, val in np.ndenumerate(peptide_layout):
            r = index[0]
            c = index[1]
            worksheet_2.cell(row=r+8, column=c+1).value = val
        # Saves workbook
        workbook.save(filename=new_csv)
